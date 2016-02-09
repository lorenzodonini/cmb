__author__ = 'lorenzodonini'

from openpyxl import Workbook

class report():
    def __init__(self):
        self.currentSheetIndex = -1
        self.currentSheet = None
        self.wb = None

    def createNewWorkbook(self):
        self.wb = Workbook()
        self.currentSheetIndex = -1
        self.currentRow = 1

    def createNewSheet(self, title):
        self.currentSheetIndex += 1
        self.currentSheet = self.wb.create_sheet(title,self.currentSheetIndex)
        self.currentRow = 1
        return self.currentSheet

    def writeHeaderLine(self, values):
        if self.currentSheet != None:
            col = 1
            for val in values:
                self.currentSheet.cell(row=1,column=col).value = val
                col += 1
        if self.currentRow == 1:
            self.currentRow += 1

    def writeLine(self, values, row=None):
        if self.currentSheet != None:
            if row == None:
                row = self.currentRow
                self.currentRow += 1
            col = 1
            for val in values:
                self.currentSheet.cell(row=row,column=col).value = val
                col += 1

    def saveWorkbook(self, filename):
        self.wb.save(filename)
